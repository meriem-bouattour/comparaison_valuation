#!/usr/bin/env python3
"""
Script de comparaison de tables Oracle et PostgreSQL
Génère un rapport Excel détaillé avec plusieurs onglets
"""

import os
import sys
from datetime import datetime
from typing import Dict, List, Tuple, Any, Optional
import logging

try:
    import cx_Oracle
    import psycopg2
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from dotenv import load_dotenv
except ImportError as e:
    print(f"Erreur d'importation: {e}")
    print("Veuillez installer les dépendances: pip install -r requirements.txt")
    sys.exit(1)

# Configuration du logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class DatabaseComparator:
    """Classe principale pour comparer les tables Oracle et PostgreSQL"""
    
    def __init__(self):
        """Initialise les connexions aux bases de données"""
        load_dotenv()
        
        # Configuration Oracle
        self.oracle_config = {
            'host': os.getenv('ORACLE_HOST'),
            'port': os.getenv('ORACLE_PORT', '1521'),
            'service_name': os.getenv('ORACLE_SERVICE_NAME'),
            'user': os.getenv('ORACLE_USER'),
            'password': os.getenv('ORACLE_PASSWORD'),
            'table': os.getenv('ORACLE_TABLE', 'VALUATION_REFPROD')
        }
        
        # Configuration PostgreSQL
        self.postgres_config = {
            'host': os.getenv('POSTGRES_HOST'),
            'port': os.getenv('POSTGRES_PORT', '5432'),
            'database': os.getenv('POSTGRES_DATABASE'),
            'user': os.getenv('POSTGRES_USER'),
            'password': os.getenv('POSTGRES_PASSWORD'),
            'schema': os.getenv('POSTGRES_SCHEMA', 'public'),
            'table': os.getenv('POSTGRES_TABLE', 'valuation_ids')
        }
        
        self.excel_output = os.getenv('EXCEL_OUTPUT_PATH', 'rapport_comparaison.xlsx')
        
        self.oracle_conn = None
        self.postgres_conn = None
        
    def connect_oracle(self) -> bool:
        """Établit la connexion à Oracle"""
        try:
            logger.info("Connexion à Oracle...")
            dsn = cx_Oracle.makedsn(
                self.oracle_config['host'],
                self.oracle_config['port'],
                service_name=self.oracle_config['service_name']
            )
            self.oracle_conn = cx_Oracle.connect(
                user=self.oracle_config['user'],
                password=self.oracle_config['password'],
                dsn=dsn,
                encoding="UTF-8"
            )
            logger.info("✓ Connexion Oracle établie")
            return True
        except cx_Oracle.Error as e:
            logger.error(f"✗ Erreur de connexion Oracle: {e}")
            return False
    
    def connect_postgres(self) -> bool:
        """Établit la connexion à PostgreSQL"""
        try:
            logger.info("Connexion à PostgreSQL...")
            self.postgres_conn = psycopg2.connect(
                host=self.postgres_config['host'],
                port=self.postgres_config['port'],
                database=self.postgres_config['database'],
                user=self.postgres_config['user'],
                password=self.postgres_config['password']
            )
            logger.info("✓ Connexion PostgreSQL établie")
            return True
        except psycopg2.Error as e:
            logger.error(f"✗ Erreur de connexion PostgreSQL: {e}")
            return False
    
    def get_oracle_columns(self) -> pd.DataFrame:
        """Récupère les métadonnées des colonnes Oracle"""
        logger.info("Extraction des colonnes Oracle...")
        
        query = """
        SELECT 
            COLUMN_NAME,
            DATA_TYPE,
            DATA_LENGTH,
            DATA_PRECISION,
            DATA_SCALE,
            NULLABLE,
            DATA_DEFAULT
        FROM ALL_TAB_COLUMNS
        WHERE TABLE_NAME = :table_name
        AND OWNER = :owner
        ORDER BY COLUMN_ID
        """
        
        cursor = self.oracle_conn.cursor()
        cursor.execute(query, {
            'table_name': self.oracle_config['table'],
            'owner': self.oracle_config['user'].upper()
        })
        
        columns = cursor.fetchall()
        df = pd.DataFrame(columns, columns=[
            'column_name', 'data_type', 'data_length', 
            'data_precision', 'data_scale', 'nullable', 'default_value'
        ])
        
        cursor.close()
        logger.info(f"✓ {len(df)} colonnes Oracle extraites")
        return df
    
    def get_postgres_columns(self) -> pd.DataFrame:
        """Récupère les métadonnées des colonnes PostgreSQL"""
        logger.info("Extraction des colonnes PostgreSQL...")
        
        query = """
        SELECT 
            column_name,
            data_type,
            character_maximum_length,
            numeric_precision,
            numeric_scale,
            is_nullable,
            column_default
        FROM information_schema.columns
        WHERE table_name = %s
        AND table_schema = %s
        ORDER BY ordinal_position
        """
        
        cursor = self.postgres_conn.cursor()
        cursor.execute(query, (
            self.postgres_config['table'],
            self.postgres_config['schema']
        ))
        
        columns = cursor.fetchall()
        df = pd.DataFrame(columns, columns=[
            'column_name', 'data_type', 'data_length',
            'data_precision', 'data_scale', 'nullable', 'default_value'
        ])
        
        cursor.close()
        logger.info(f"✓ {len(df)} colonnes PostgreSQL extraites")
        return df
    
    def get_oracle_constraints(self) -> Dict[str, pd.DataFrame]:
        """Récupère les contraintes Oracle"""
        logger.info("Extraction des contraintes Oracle...")
        
        result = {}
        
        # Clés primaires
        pk_query = """
        SELECT 
            cols.column_name,
            cons.constraint_name
        FROM all_constraints cons
        JOIN all_cons_columns cols ON cons.constraint_name = cols.constraint_name
        WHERE cons.table_name = :table_name
        AND cons.owner = :owner
        AND cons.constraint_type = 'P'
        ORDER BY cols.position
        """
        
        cursor = self.oracle_conn.cursor()
        cursor.execute(pk_query, {
            'table_name': self.oracle_config['table'],
            'owner': self.oracle_config['user'].upper()
        })
        result['primary_keys'] = pd.DataFrame(
            cursor.fetchall(),
            columns=['column_name', 'constraint_name']
        )
        
        # Clés étrangères
        fk_query = """
        SELECT 
            cols.column_name,
            cons.constraint_name,
            cons.r_constraint_name as referenced_constraint
        FROM all_constraints cons
        JOIN all_cons_columns cols ON cons.constraint_name = cols.constraint_name
        WHERE cons.table_name = :table_name
        AND cons.owner = :owner
        AND cons.constraint_type = 'R'
        ORDER BY cols.position
        """
        
        cursor.execute(fk_query, {
            'table_name': self.oracle_config['table'],
            'owner': self.oracle_config['user'].upper()
        })
        result['foreign_keys'] = pd.DataFrame(
            cursor.fetchall(),
            columns=['column_name', 'constraint_name', 'referenced_constraint']
        )
        
        # Contraintes UNIQUE
        uk_query = """
        SELECT 
            cols.column_name,
            cons.constraint_name
        FROM all_constraints cons
        JOIN all_cons_columns cols ON cons.constraint_name = cols.constraint_name
        WHERE cons.table_name = :table_name
        AND cons.owner = :owner
        AND cons.constraint_type = 'U'
        ORDER BY cols.position
        """
        
        cursor.execute(uk_query, {
            'table_name': self.oracle_config['table'],
            'owner': self.oracle_config['user'].upper()
        })
        result['unique_constraints'] = pd.DataFrame(
            cursor.fetchall(),
            columns=['column_name', 'constraint_name']
        )
        
        cursor.close()
        logger.info("✓ Contraintes Oracle extraites")
        return result
    
    def get_postgres_constraints(self) -> Dict[str, pd.DataFrame]:
        """Récupère les contraintes PostgreSQL"""
        logger.info("Extraction des contraintes PostgreSQL...")
        
        result = {}
        
        # Clés primaires
        pk_query = """
        SELECT 
            kcu.column_name,
            tc.constraint_name
        FROM information_schema.table_constraints tc
        JOIN information_schema.key_column_usage kcu 
            ON tc.constraint_name = kcu.constraint_name
        WHERE tc.table_name = %s
        AND tc.table_schema = %s
        AND tc.constraint_type = 'PRIMARY KEY'
        ORDER BY kcu.ordinal_position
        """
        
        cursor = self.postgres_conn.cursor()
        cursor.execute(pk_query, (
            self.postgres_config['table'],
            self.postgres_config['schema']
        ))
        result['primary_keys'] = pd.DataFrame(
            cursor.fetchall(),
            columns=['column_name', 'constraint_name']
        )
        
        # Clés étrangères
        fk_query = """
        SELECT 
            kcu.column_name,
            tc.constraint_name,
            ccu.table_name as referenced_table,
            ccu.column_name as referenced_column
        FROM information_schema.table_constraints tc
        JOIN information_schema.key_column_usage kcu 
            ON tc.constraint_name = kcu.constraint_name
        JOIN information_schema.constraint_column_usage ccu
            ON tc.constraint_name = ccu.constraint_name
        WHERE tc.table_name = %s
        AND tc.table_schema = %s
        AND tc.constraint_type = 'FOREIGN KEY'
        ORDER BY kcu.ordinal_position
        """
        
        cursor.execute(fk_query, (
            self.postgres_config['table'],
            self.postgres_config['schema']
        ))
        result['foreign_keys'] = pd.DataFrame(
            cursor.fetchall(),
            columns=['column_name', 'constraint_name', 'referenced_table', 'referenced_column']
        )
        
        # Contraintes UNIQUE
        uk_query = """
        SELECT 
            kcu.column_name,
            tc.constraint_name
        FROM information_schema.table_constraints tc
        JOIN information_schema.key_column_usage kcu 
            ON tc.constraint_name = kcu.constraint_name
        WHERE tc.table_name = %s
        AND tc.table_schema = %s
        AND tc.constraint_type = 'UNIQUE'
        ORDER BY kcu.ordinal_position
        """
        
        cursor.execute(uk_query, (
            self.postgres_config['table'],
            self.postgres_config['schema']
        ))
        result['unique_constraints'] = pd.DataFrame(
            cursor.fetchall(),
            columns=['column_name', 'constraint_name']
        )
        
        cursor.close()
        logger.info("✓ Contraintes PostgreSQL extraites")
        return result
    
    def get_oracle_indexes(self) -> pd.DataFrame:
        """Récupère les index Oracle"""
        logger.info("Extraction des index Oracle...")
        
        query = """
        SELECT 
            i.index_name,
            ic.column_name,
            i.uniqueness,
            i.index_type,
            ic.column_position
        FROM all_indexes i
        JOIN all_ind_columns ic ON i.index_name = ic.index_name
        WHERE i.table_name = :table_name
        AND i.owner = :owner
        ORDER BY i.index_name, ic.column_position
        """
        
        cursor = self.oracle_conn.cursor()
        cursor.execute(query, {
            'table_name': self.oracle_config['table'],
            'owner': self.oracle_config['user'].upper()
        })
        
        df = pd.DataFrame(
            cursor.fetchall(),
            columns=['index_name', 'column_name', 'uniqueness', 'index_type', 'column_position']
        )
        
        cursor.close()
        logger.info(f"✓ {len(df)} index Oracle extraits")
        return df
    
    def get_postgres_indexes(self) -> pd.DataFrame:
        """Récupère les index PostgreSQL"""
        logger.info("Extraction des index PostgreSQL...")
        
        query = """
        SELECT 
            i.indexname as index_name,
            a.attname as column_name,
            ix.indisunique as is_unique,
            am.amname as index_type
        FROM pg_indexes i
        JOIN pg_class t ON i.tablename = t.relname
        JOIN pg_index ix ON t.oid = ix.indrelid
        JOIN pg_class idx ON ix.indexrelid = idx.oid
        JOIN pg_am am ON idx.relam = am.oid
        JOIN pg_attribute a ON a.attrelid = t.oid AND a.attnum = ANY(ix.indkey)
        WHERE i.tablename = %s
        AND i.schemaname = %s
        ORDER BY i.indexname, a.attnum
        """
        
        cursor = self.postgres_conn.cursor()
        cursor.execute(query, (
            self.postgres_config['table'],
            self.postgres_config['schema']
        ))
        
        df = pd.DataFrame(
            cursor.fetchall(),
            columns=['index_name', 'column_name', 'is_unique', 'index_type']
        )
        
        cursor.close()
        logger.info(f"✓ {len(df)} index PostgreSQL extraits")
        return df
    
    def get_oracle_statistics(self, columns_df: pd.DataFrame) -> pd.DataFrame:
        """Récupère les statistiques Oracle"""
        logger.info("Calcul des statistiques Oracle...")
        
        cursor = self.oracle_conn.cursor()
        
        # Nombre total de lignes
        query = f"SELECT COUNT(*) FROM {self.oracle_config['table']}"
        cursor.execute(query)
        total_rows = cursor.fetchone()[0]
        
        stats = []
        for _, col in columns_df.iterrows():
            col_name = col['column_name']
            col_type = col['data_type']
            
            # Statistiques de base
            null_query = f"""
            SELECT 
                COUNT(*) - COUNT({col_name}) as null_count,
                COUNT(DISTINCT {col_name}) as distinct_count
            FROM {self.oracle_config['table']}
            """
            
            try:
                cursor.execute(null_query)
                null_count, distinct_count = cursor.fetchone()
                
                stat = {
                    'column_name': col_name,
                    'total_rows': total_rows,
                    'null_count': null_count,
                    'distinct_count': distinct_count
                }
                
                # Statistiques pour colonnes numériques
                if col_type in ['NUMBER', 'INTEGER', 'FLOAT']:
                    num_query = f"""
                    SELECT 
                        MIN({col_name}),
                        MAX({col_name}),
                        AVG({col_name})
                    FROM {self.oracle_config['table']}
                    """
                    cursor.execute(num_query)
                    min_val, max_val, avg_val = cursor.fetchone()
                    stat.update({
                        'min_value': min_val,
                        'max_value': max_val,
                        'avg_value': avg_val
                    })
                
                # Statistiques pour colonnes texte
                elif col_type in ['VARCHAR2', 'CHAR', 'CLOB']:
                    len_query = f"""
                    SELECT 
                        MIN(LENGTH({col_name})),
                        MAX(LENGTH({col_name}))
                    FROM {self.oracle_config['table']}
                    WHERE {col_name} IS NOT NULL
                    """
                    cursor.execute(len_query)
                    result = cursor.fetchone()
                    if result:
                        stat.update({
                            'min_length': result[0],
                            'max_length': result[1]
                        })
                
                stats.append(stat)
                
            except Exception as e:
                logger.warning(f"Impossible de calculer les statistiques pour {col_name}: {e}")
                continue
        
        cursor.close()
        logger.info(f"✓ Statistiques calculées pour {len(stats)} colonnes Oracle")
        return pd.DataFrame(stats)
    
    def get_postgres_statistics(self, columns_df: pd.DataFrame) -> pd.DataFrame:
        """Récupère les statistiques PostgreSQL"""
        logger.info("Calcul des statistiques PostgreSQL...")
        
        cursor = self.postgres_conn.cursor()
        
        # Nombre total de lignes
        query = f"SELECT COUNT(*) FROM {self.postgres_config['schema']}.{self.postgres_config['table']}"
        cursor.execute(query)
        total_rows = cursor.fetchone()[0]
        
        stats = []
        for _, col in columns_df.iterrows():
            col_name = col['column_name']
            col_type = col['data_type']
            
            # Statistiques de base
            null_query = f"""
            SELECT 
                COUNT(*) - COUNT({col_name}) as null_count,
                COUNT(DISTINCT {col_name}) as distinct_count
            FROM {self.postgres_config['schema']}.{self.postgres_config['table']}
            """
            
            try:
                cursor.execute(null_query)
                null_count, distinct_count = cursor.fetchone()
                
                stat = {
                    'column_name': col_name,
                    'total_rows': total_rows,
                    'null_count': null_count,
                    'distinct_count': distinct_count
                }
                
                # Statistiques pour colonnes numériques
                if col_type in ['integer', 'bigint', 'numeric', 'real', 'double precision']:
                    num_query = f"""
                    SELECT 
                        MIN({col_name}),
                        MAX({col_name}),
                        AVG({col_name})
                    FROM {self.postgres_config['schema']}.{self.postgres_config['table']}
                    """
                    cursor.execute(num_query)
                    min_val, max_val, avg_val = cursor.fetchone()
                    stat.update({
                        'min_value': min_val,
                        'max_value': max_val,
                        'avg_value': avg_val
                    })
                
                # Statistiques pour colonnes texte
                elif col_type in ['character varying', 'character', 'text']:
                    len_query = f"""
                    SELECT 
                        MIN(LENGTH({col_name})),
                        MAX(LENGTH({col_name}))
                    FROM {self.postgres_config['schema']}.{self.postgres_config['table']}
                    WHERE {col_name} IS NOT NULL
                    """
                    cursor.execute(len_query)
                    result = cursor.fetchone()
                    if result:
                        stat.update({
                            'min_length': result[0],
                            'max_length': result[1]
                        })
                
                stats.append(stat)
                
            except Exception as e:
                logger.warning(f"Impossible de calculer les statistiques pour {col_name}: {e}")
                continue
        
        cursor.close()
        logger.info(f"✓ Statistiques calculées pour {len(stats)} colonnes PostgreSQL")
        return pd.DataFrame(stats)
    
    def compare_columns(self, oracle_cols: pd.DataFrame, postgres_cols: pd.DataFrame) -> pd.DataFrame:
        """Compare les colonnes des deux tables"""
        logger.info("Comparaison des colonnes...")
        
        oracle_names = set(oracle_cols['column_name'].str.lower())
        postgres_names = set(postgres_cols['column_name'].str.lower())
        
        only_oracle = oracle_names - postgres_names
        only_postgres = postgres_names - oracle_names
        common = oracle_names & postgres_names
        
        comparison = []
        
        # Colonnes uniquement dans Oracle
        for col in only_oracle:
            comparison.append({
                'column_name': col,
                'in_oracle': 'Oui',
                'in_postgres': 'Non',
                'status': 'Oracle uniquement'
            })
        
        # Colonnes uniquement dans PostgreSQL
        for col in only_postgres:
            comparison.append({
                'column_name': col,
                'in_oracle': 'Non',
                'in_postgres': 'Oui',
                'status': 'PostgreSQL uniquement'
            })
        
        # Colonnes communes
        for col in common:
            comparison.append({
                'column_name': col,
                'in_oracle': 'Oui',
                'in_postgres': 'Oui',
                'status': 'Commun'
            })
        
        df = pd.DataFrame(comparison)
        logger.info(f"✓ Comparaison terminée: {len(common)} colonnes communes, "
                   f"{len(only_oracle)} uniquement Oracle, {len(only_postgres)} uniquement PostgreSQL")
        return df
    
    def compare_data_types(self, oracle_cols: pd.DataFrame, postgres_cols: pd.DataFrame) -> pd.DataFrame:
        """Compare les types de données des colonnes communes"""
        logger.info("Comparaison des types de données...")
        
        # Créer des dictionnaires pour faciliter la recherche
        oracle_dict = {row['column_name'].lower(): row for _, row in oracle_cols.iterrows()}
        postgres_dict = {row['column_name'].lower(): row for _, row in postgres_cols.iterrows()}
        
        common_cols = set(oracle_dict.keys()) & set(postgres_dict.keys())
        
        comparison = []
        for col_name in common_cols:
            oracle_info = oracle_dict[col_name]
            postgres_info = postgres_dict[col_name]
            
            # Construire la description du type
            oracle_type_desc = oracle_info['data_type']
            if oracle_info['data_precision']:
                oracle_type_desc += f"({oracle_info['data_precision']}"
                if oracle_info['data_scale']:
                    oracle_type_desc += f",{oracle_info['data_scale']}"
                oracle_type_desc += ")"
            elif oracle_info['data_length']:
                oracle_type_desc += f"({oracle_info['data_length']})"
            
            postgres_type_desc = postgres_info['data_type']
            if postgres_info['data_precision']:
                postgres_type_desc += f"({postgres_info['data_precision']}"
                if postgres_info['data_scale']:
                    postgres_type_desc += f",{postgres_info['data_scale']}"
                postgres_type_desc += ")"
            elif postgres_info['data_length']:
                postgres_type_desc += f"({postgres_info['data_length']})"
            
            # Comparer
            types_match = self._are_types_compatible(
                oracle_info['data_type'],
                postgres_info['data_type']
            )
            
            comparison.append({
                'column_name': col_name,
                'oracle_type': oracle_type_desc,
                'postgres_type': postgres_type_desc,
                'compatible': 'Oui' if types_match else 'Non',
                'oracle_nullable': oracle_info['nullable'],
                'postgres_nullable': postgres_info['nullable']
            })
        
        df = pd.DataFrame(comparison)
        logger.info(f"✓ Types de données comparés pour {len(df)} colonnes")
        return df
    
    @staticmethod
    def _are_types_compatible(oracle_type: str, postgres_type: str) -> bool:
        """Vérifie si les types Oracle et PostgreSQL sont compatibles"""
        type_mapping = {
            'NUMBER': ['numeric', 'integer', 'bigint', 'real', 'double precision'],
            'INTEGER': ['integer', 'bigint'],
            'VARCHAR2': ['character varying', 'text'],
            'CHAR': ['character', 'character varying'],
            'DATE': ['date', 'timestamp without time zone', 'timestamp with time zone'],
            'TIMESTAMP': ['timestamp without time zone', 'timestamp with time zone'],
            'CLOB': ['text'],
            'BLOB': ['bytea']
        }
        
        if oracle_type in type_mapping:
            return postgres_type in type_mapping[oracle_type]
        
        return oracle_type.lower() == postgres_type.lower()
    
    def generate_summary(self, all_data: Dict) -> pd.DataFrame:
        """Génère le résumé de la comparaison"""
        logger.info("Génération du résumé...")
        
        oracle_cols = all_data['oracle_columns']
        postgres_cols = all_data['postgres_columns']
        columns_comparison = all_data['columns_comparison']
        
        common_cols = len(columns_comparison[columns_comparison['status'] == 'Commun'])
        total_cols = len(columns_comparison)
        similarity = (common_cols / total_cols * 100) if total_cols > 0 else 0
        
        summary_data = [
            ['Date de génération', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['', ''],
            ['=== COLONNES ===', ''],
            ['Colonnes Oracle', len(oracle_cols)],
            ['Colonnes PostgreSQL', len(postgres_cols)],
            ['Colonnes communes', common_cols],
            ['Colonnes uniquement Oracle', len(columns_comparison[columns_comparison['status'] == 'Oracle uniquement'])],
            ['Colonnes uniquement PostgreSQL', len(columns_comparison[columns_comparison['status'] == 'PostgreSQL uniquement'])],
            ['', ''],
            ['=== ENREGISTREMENTS ===', ''],
        ]
        
        # Ajouter les stats si disponibles
        if 'oracle_statistics' in all_data and not all_data['oracle_statistics'].empty:
            oracle_rows = all_data['oracle_statistics'].iloc[0]['total_rows']
            summary_data.append(['Enregistrements Oracle', oracle_rows])
        
        if 'postgres_statistics' in all_data and not all_data['postgres_statistics'].empty:
            postgres_rows = all_data['postgres_statistics'].iloc[0]['total_rows']
            summary_data.append(['Enregistrements PostgreSQL', postgres_rows])
        
        summary_data.extend([
            ['', ''],
            ['=== CONTRAINTES ===', ''],
            ['Clés primaires Oracle', len(all_data['oracle_constraints']['primary_keys'])],
            ['Clés primaires PostgreSQL', len(all_data['postgres_constraints']['primary_keys'])],
            ['Clés étrangères Oracle', len(all_data['oracle_constraints']['foreign_keys'])],
            ['Clés étrangères PostgreSQL', len(all_data['postgres_constraints']['foreign_keys'])],
            ['', ''],
            ['=== INDEX ===', ''],
            ['Index Oracle', len(all_data['oracle_indexes']['index_name'].unique()) if not all_data['oracle_indexes'].empty else 0],
            ['Index PostgreSQL', len(all_data['postgres_indexes']['index_name'].unique()) if not all_data['postgres_indexes'].empty else 0],
            ['', ''],
            ['=== SCORE ===', ''],
            ['Score de similarité', f'{similarity:.2f}%'],
        ])
        
        df = pd.DataFrame(summary_data, columns=['Métrique', 'Valeur'])
        logger.info("✓ Résumé généré")
        return df
    
    def format_excel(self, excel_file: str):
        """Applique le formatage au fichier Excel"""
        logger.info("Formatage du fichier Excel...")
        
        wb = load_workbook(excel_file)
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Formater l'en-tête
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Coloration conditionnelle selon le contenu
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.value:
                        value_str = str(cell.value).lower()
                        
                        # Vert pour "Oui", "Commun", etc.
                        if value_str in ['oui', 'commun', 'yes']:
                            cell.fill = green_fill
                        
                        # Rouge pour "Non", "uniquement", etc.
                        elif value_str in ['non', 'no'] or 'uniquement' in value_str:
                            cell.fill = red_fill
                        
                        # Jaune pour "attention", "warning"
                        elif 'attention' in value_str or 'warning' in value_str:
                            cell.fill = yellow_fill
            
            # Auto-dimensionner les colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Figer la première ligne
            ws.freeze_panes = 'A2'
        
        wb.save(excel_file)
        logger.info("✓ Formatage terminé")
    
    def generate_report(self):
        """Génère le rapport Excel complet"""
        logger.info("="*50)
        logger.info("DÉBUT DE LA COMPARAISON DES TABLES")
        logger.info("="*50)
        
        # Connexions
        if not self.connect_oracle():
            logger.error("Impossible de se connecter à Oracle. Abandon.")
            return False
        
        if not self.connect_postgres():
            logger.error("Impossible de se connecter à PostgreSQL. Abandon.")
            return False
        
        try:
            # Collecte des données
            logger.info("\n--- Extraction des métadonnées ---")
            
            oracle_cols = self.get_oracle_columns()
            postgres_cols = self.get_postgres_columns()
            
            oracle_constraints = self.get_oracle_constraints()
            postgres_constraints = self.get_postgres_constraints()
            
            oracle_indexes = self.get_oracle_indexes()
            postgres_indexes = self.get_postgres_indexes()
            
            # Comparaisons
            logger.info("\n--- Comparaisons ---")
            
            columns_comparison = self.compare_columns(oracle_cols, postgres_cols)
            data_types_comparison = self.compare_data_types(oracle_cols, postgres_cols)
            
            # Statistiques (optionnel - peut être lent)
            logger.info("\n--- Statistiques ---")
            try:
                oracle_stats = self.get_oracle_statistics(oracle_cols)
                postgres_stats = self.get_postgres_statistics(postgres_cols)
            except Exception as e:
                logger.warning(f"Impossible de calculer les statistiques: {e}")
                oracle_stats = pd.DataFrame()
                postgres_stats = pd.DataFrame()
            
            # Assemblage des données
            all_data = {
                'oracle_columns': oracle_cols,
                'postgres_columns': postgres_cols,
                'oracle_constraints': oracle_constraints,
                'postgres_constraints': postgres_constraints,
                'oracle_indexes': oracle_indexes,
                'postgres_indexes': postgres_indexes,
                'columns_comparison': columns_comparison,
                'data_types_comparison': data_types_comparison,
                'oracle_statistics': oracle_stats,
                'postgres_statistics': postgres_stats
            }
            
            summary = self.generate_summary(all_data)
            
            # Génération du fichier Excel
            logger.info("\n--- Génération du rapport Excel ---")
            
            with pd.ExcelWriter(self.excel_output, engine='openpyxl') as writer:
                # Onglet Résumé
                summary.to_excel(writer, sheet_name='Résumé', index=False)
                
                # Onglet Colonnes
                columns_comparison.to_excel(writer, sheet_name='Colonnes', index=False)
                
                # Onglet Types de données
                data_types_comparison.to_excel(writer, sheet_name='Types de données', index=False)
                
                # Onglet Contraintes
                constraints_df = pd.DataFrame([
                    ['=== CLÉS PRIMAIRES ORACLE ===', '', ''],
                ])
                constraints_df = pd.concat([
                    constraints_df,
                    oracle_constraints['primary_keys'].rename(columns={
                        'column_name': 'Colonne',
                        'constraint_name': 'Nom contrainte'
                    }).assign(Type='PK Oracle')
                ], ignore_index=True)
                
                constraints_df = pd.concat([
                    constraints_df,
                    pd.DataFrame([['', '', ''], ['=== CLÉS PRIMAIRES POSTGRESQL ===', '', '']]),
                    postgres_constraints['primary_keys'].rename(columns={
                        'column_name': 'Colonne',
                        'constraint_name': 'Nom contrainte'
                    }).assign(Type='PK PostgreSQL')
                ], ignore_index=True)
                
                constraints_df = pd.concat([
                    constraints_df,
                    pd.DataFrame([['', '', ''], ['=== CLÉS ÉTRANGÈRES ORACLE ===', '', '']]),
                    oracle_constraints['foreign_keys'].rename(columns={
                        'column_name': 'Colonne',
                        'constraint_name': 'Nom contrainte'
                    }).assign(Type='FK Oracle')
                ], ignore_index=True)
                
                constraints_df = pd.concat([
                    constraints_df,
                    pd.DataFrame([['', '', ''], ['=== CLÉS ÉTRANGÈRES POSTGRESQL ===', '', '']]),
                    postgres_constraints['foreign_keys'][['column_name', 'constraint_name']].rename(columns={
                        'column_name': 'Colonne',
                        'constraint_name': 'Nom contrainte'
                    }).assign(Type='FK PostgreSQL')
                ], ignore_index=True)
                
                constraints_df.to_excel(writer, sheet_name='Clés et Contraintes', index=False, header=False)
                
                # Onglet Index
                indexes_df = pd.DataFrame([['=== INDEX ORACLE ===', '', '', '']])
                indexes_df = pd.concat([
                    indexes_df,
                    oracle_indexes.rename(columns={
                        'index_name': 'Nom index',
                        'column_name': 'Colonne',
                        'uniqueness': 'Unique',
                        'index_type': 'Type'
                    })[['Nom index', 'Colonne', 'Unique', 'Type']]
                ], ignore_index=True)
                
                indexes_df = pd.concat([
                    indexes_df,
                    pd.DataFrame([['', '', '', ''], ['=== INDEX POSTGRESQL ===', '', '', '']]),
                    postgres_indexes.rename(columns={
                        'index_name': 'Nom index',
                        'column_name': 'Colonne',
                        'is_unique': 'Unique',
                        'index_type': 'Type'
                    })[['Nom index', 'Colonne', 'Unique', 'Type']]
                ], ignore_index=True)
                
                indexes_df.to_excel(writer, sheet_name='Index', index=False, header=False)
                
                # Onglet Statistiques
                if not oracle_stats.empty or not postgres_stats.empty:
                    stats_df = pd.DataFrame([['=== STATISTIQUES ORACLE ==='] + ['']*(len(oracle_stats.columns)-1)])
                    
                    if not oracle_stats.empty:
                        stats_df = pd.concat([stats_df, oracle_stats], ignore_index=True)
                    
                    stats_df = pd.concat([
                        stats_df,
                        pd.DataFrame([[''] * len(oracle_stats.columns) if not oracle_stats.empty else [''],
                                     ['=== STATISTIQUES POSTGRESQL ==='] + ['']*(len(postgres_stats.columns)-1 if not postgres_stats.empty else 0)])
                    ], ignore_index=True)
                    
                    if not postgres_stats.empty:
                        stats_df = pd.concat([stats_df, postgres_stats], ignore_index=True)
                    
                    stats_df.to_excel(writer, sheet_name='Statistiques', index=False, header=False)
            
            # Formatage
            self.format_excel(self.excel_output)
            
            logger.info(f"\n✓ Rapport généré avec succès: {self.excel_output}")
            logger.info("="*50)
            logger.info("COMPARAISON TERMINÉE")
            logger.info("="*50)
            
            return True
            
        except Exception as e:
            logger.error(f"Erreur lors de la génération du rapport: {e}", exc_info=True)
            return False
        
        finally:
            # Fermeture des connexions
            if self.oracle_conn:
                self.oracle_conn.close()
                logger.info("Connexion Oracle fermée")
            
            if self.postgres_conn:
                self.postgres_conn.close()
                logger.info("Connexion PostgreSQL fermée")


def main():
    """Point d'entrée principal"""
    print("""
╔═══════════════════════════════════════════════════════════════╗
║   COMPARATEUR DE TABLES ORACLE ↔ POSTGRESQL                  ║
║   Génère un rapport Excel détaillé                            ║
╚═══════════════════════════════════════════════════════════════╝
    """)
    
    # Vérifier que le fichier .env existe
    if not os.path.exists('.env'):
        logger.warning("Fichier .env non trouvé. Veuillez créer un fichier .env basé sur .env.example")
        logger.warning("Les variables d'environnement système seront utilisées.")
    
    comparator = DatabaseComparator()
    success = comparator.generate_report()
    
    if success:
        print(f"\n✓ Rapport Excel généré: {comparator.excel_output}")
        return 0
    else:
        print("\n✗ Échec de la génération du rapport")
        return 1


if __name__ == '__main__':
    sys.exit(main())
